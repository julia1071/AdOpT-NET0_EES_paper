import re
import pandas as pd
import os

from openpyxl.reader.excel import load_workbook

execute = 0

if execute:
    # Load the Excel file
    file_path = "C:/Users/5637635/PycharmProjects/AdOpT-NET0_Julia/Plotting/result_data_long.xlsx"
    df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)

    # Use the first and third rows as headers
    top_header = df.iloc[0].ffill()  # Forward fill merged cells
    top_header = top_header.replace({"EmissionLimit Greenfield": "Greenfield (Scope 1, 2, and 3)", "EmissionLimit Brownfield": "Brownfield (Scope 1, 2, and 3)",
                                     "EmissionScope Greenfield": "Greenfield (Scope 1 and 2)", "EmissionScope Brownfield":
                                         "Brownfield (Scope 1 and 2)",})
    sub_header = df.iloc[2].replace({"2030": "Short-term", "2040": "Mid-term", "2050": "Long-term"})

    # sub_header = df.iloc[2].astype(str)  # Convert to string
    df.columns = pd.MultiIndex.from_arrays([top_header, sub_header])

    # Rename first column header
    top_header = top_header.copy()  # Avoid modifying the original header list
    top_header.iloc[0] = ""
    sub_header.iloc[0] = "Result type"
    df = df.iloc[4:].reset_index(drop=True)

    # Define the filter criteria
    keywords = [
        "size_"
    ]
    keywords_import = [
        "electricity/import_max", "CO2/export_max",
        "methane/import_max", "methane_bio/import_max",
        "CO2_DAC/import_max", "MPW/import_max", "propane/import_max"
    ]

    # Extract first column name
    first_column_name = df.columns[0]
    first_column_data = df[first_column_name].squeeze()

    # Filter rows based on keywords
    filtered_df = df[df[first_column_name].astype(str).str.startswith(tuple(keywords))]
    filtered_df_import = df[first_column_data.astype(str).str.startswith(tuple(keywords_import))]

    # Remove unwanted rows
    filtered_df = filtered_df[
        ~filtered_df[first_column_name].str.contains("mixer|CO2toEmission|WGS|OlefinSeparation", na=False, case=False)]

    name_unit_mapping = {
        "size_CrackerFurnace": (r"Conventional cracker", "t naphtha/h"),
        "size_CrackerFurnace_CC": (r"Conventional cracker with \acs{CC}", "\% captured"),
        "size_CrackerFurnace_Electric": (r"Electric cracker", "t naphtha/h"),
        "size_SteamReformer": (r"Conventional reformer", "MW gas"),
        "size_SteamReformer_CC": (r"Conventional reformer with \acs{CC}", "\% captured"),
        "size_ElectricSMR_m": (r"Electric reformer", "MW gas"),
        "size_AEC": (r"\acs{AEC}", "MW electric"),
        "size_HaberBosch": (r"\acs{HB} process", "MW hydrogen"),
        "size_RWGS": ("RWGS", "t \ce{CO2}/h"),
        "size_MeOHsynthesis": (r"Methanol synthesis from syngas", "t syngas/h"),
        "size_DirectMeOHsynthesis": (r"Direct methanol synthesis from \ce{CO2}", "t \ce{CO2}/h"),
        "size_MTO": (r"\acs{MTO}", "t methanol/h"),
        "size_EDH": (r"\acs{EDH}", "t ethanol/h"),
        "size_PDH": (r"\acs{PDH}", "t propane/h"),
        "size_MPW2methanol": (r"\acs{MPW}-to-methanol", "t MPW/h"),
        "size_MPW2methanol_CC": (r"\acs{MPW}-to-methanol with \acs{CC}", "\% captured"),
        "size_CO2electrolysis": (r"\ce{CO2} electrolysis", "t \ce{CO2}/h"),
        "size_ASU": (r"\acs{ASU}", "MW electricity"),
        "size_Boiler_Industrial_NG": (r"Gas-fired boiler", "MW gas"),
        "size_Boiler_El": (r"Electric boiler", "MW electricity"),
        # Storage & Import Entries (Keep These at the Bottom)
        "size_Storage_Ammonia": ("Ammonia tank", "tonne"),
        "size_Storage_Battery": ("Li-ion battery", "MWh"),
        "size_Storage_CO2": (r"\ce{CO2} buffer storage", "tonne"),
        "size_Storage_H2": ("Hydrogen tank", "MWh"),
        "size_Storage_Ethylene": ("Ethylene tank", "tonne"),
        "size_Storage_Propylene": ("Propylene tank", "tonne"),
        # "electricity/import_max": ("Electricity grid import", "MW (\% of max)"),
        # "CO2/export_max": (r"\ce{CO2} T\&S", "t \ce{CO2}/h (\% of max)"),
        # "MPW/import_max": ("MPW import", "t MPW/h (\% of max)"),
        # "methane/import_max": ("Methane import", "MW gas"),
        # "methane_bio/import_max": ("Bio-methane import", "MW gas"),
        # "CO2_DAC/import_max": ("DAC-\ce{CO2} import", "t \ce{CO2}/h"),
        # "propane/import_max": ("Bio-propane import", "MW"),
    }

    name_unit_mapping_import = {
        "electricity/import_max": ("Electricity grid import", "MW (\% of max)"),
        "CO2/export_max": (r"\ce{CO2} T\&S", "t \ce{CO2}/h (\% of max)"),
        "MPW/import_max": ("MPW import", "t MPW/h (\% of max)"),
        "methane/import_max": ("Methane import", "MW gas"),
        "methane_bio/import_max": ("Bio-methane import", "MW gas"),
        "CO2_DAC/import_max": ("DAC-\ce{CO2} import", "t \ce{CO2}/h"),
        "propane/import_max": ("Bio-propane import", "MW"),
    }


    # Function to rename technologies and format existing ones correctly
    def rename_tech(name):
        base_name = name.replace("_existing", "")
        tech_name, unit = name_unit_mapping.get(base_name, (name, "-"))  # Default to '-' if not found
        if "_existing" in name:
            tech_name += " existing"
        return tech_name, unit

    def rename_import(name):
        import_name, unit = name_unit_mapping_import.get(name, (name, "-"))  # Default to '-' if not found
        return import_name, unit

    # Apply renaming and extract units using .map() for better performance
    filtered_df[('', 'Technology')] = filtered_df[first_column_name].map(lambda x: rename_tech(x)[0])
    filtered_df[('', 'Unit')] = filtered_df[first_column_name].map(lambda x: rename_tech(x)[1])
    filtered_df_import[( '', 'Carrier')] = filtered_df_import[first_column_name].map(
        lambda x: rename_import(x)[0])
    filtered_df_import[('', 'Unit')] = filtered_df_import[first_column_name].map(lambda x: rename_import(x)[1])

    # Define order based on LaTeX table
    ordered_techs = list(name_unit_mapping.keys())
    ordered_imports = list(name_unit_mapping_import.keys())

    # Generate a new list of technologies with "existing" tech placed right below the original technology
    ordered_techs_with_existing = []
    for tech in ordered_techs:
        if "CC" not in tech:
            ordered_techs_with_existing.append(tech)
            if f"{tech}_CC" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_CC")
            if f"{tech}_existing" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_existing")
            if f"{tech}_existing_CC" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_existing_CC")

    # Create sorting index to maintain the correct order (with "existing" techs below the original ones)
    filtered_df['Order'] = filtered_df[first_column_name].apply(
        lambda x: ordered_techs_with_existing.index(x) if x in ordered_techs_with_existing else len(
            ordered_techs_with_existing))
    filtered_df_import['Order'] = first_column_data.apply(
        lambda x: ordered_imports.index(x) if x in name_unit_mapping_import else len(
            name_unit_mapping_import))

    # Sort the dataframe based on the 'Order' column and remove it
    filtered_df = filtered_df.sort_values(by=['Order']).drop(columns=['Order'])
    filtered_df_import = filtered_df_import.sort_values(by=['Order']).drop(columns=['Order'])

    # Round all numeric columns to whole numbers and format nan
    def custom_format_row(row):
        if 'Resulttype' in row and 'CC' in str(row['Resulttype']):
            # Preserve floats (just replace NaN with '-')
            return row.apply(lambda x: '-' if pd.isna(x) else int(x*100) if isinstance(x, float) else x)
        else:
            # Round floats (except NaN) to int
            return row.apply(lambda x: '-' if pd.isna(x) else int(round(x, 0)) if isinstance(x, float) else x)


    filtered_df = filtered_df.apply(custom_format_row, axis=1)
    filtered_df_import = filtered_df_import.apply(custom_format_row, axis=1)

    # If the index is multi-level, you can access the levels
    ordered_columns = [('', 'Technology') , ('', 'Unit')] + [col for col in filtered_df.columns if col not in [('', 'Technology'), ('', 'Unit')]]
    ordered_columns_import = [('', 'Carrier') , ('', 'Unit')] + [col for col in filtered_df.columns if col not in [('', 'Carrier'), ('', 'Unit')]]
    filtered_df = filtered_df.drop(columns=('Resulttype', 'Interval'))
    filtered_df_import = filtered_df_import.drop(columns=('Resulttype', 'Interval'))


    def swap_cc_existing(tech_str):
        if isinstance(tech_str, str):
            return re.sub(r'with (\\acs\{CC\}) existing', r'existing with \1', tech_str)
        return tech_str

    # Apply to the relevant column
    filtered_df[('', 'Technology')] = filtered_df[('', 'Technology')].apply(swap_cc_existing)

    #Import tight data
    df_tight = pd.read_pickle("df_tight.pkl")
    df_tight.columns = pd.MultiIndex.from_tuples([
        ('Brownfield Tight Emission Limit' if 'Brownfield' in col[0] else col[0], col[1])
        for col in df_tight.columns
    ])

    # Create separate DataFrames for emission limits and emission scopes
    df_emission_limit = filtered_df.loc[:, filtered_df.columns.get_level_values(0).str.contains("Scope 1, 2, and 3", na=False)]
    df_emission_scope = filtered_df.loc[:, filtered_df.columns.get_level_values(0).str.contains("Scope 1 and 2", na=False)]
    df_emission_limit = pd.concat([filtered_df[[('', 'Technology'), ('', 'Unit')]], df_emission_limit], axis=1)
    df_emission_scope = pd.concat([filtered_df[[('', 'Technology'), ('', 'Unit')]], df_emission_scope], axis=1)

    #Add tight to emission limit
    df_tight = df_tight.set_index(('', 'Technology'))
    df_tight = df_tight.drop(columns=[('', 'Unit')])
    df_emission_limit = df_emission_limit.set_index(('', 'Technology'))
    df_tight_aligned = df_tight.reindex(df_emission_limit.index)
    df_emission_limit = pd.concat([df_emission_limit, df_tight_aligned], axis=1)
    df_emission_limit = df_emission_limit.reset_index()
    df_emission_limit = df_emission_limit.apply(custom_format_row, axis=1)


    #import tables
    df_emission_limit_import = filtered_df_import.loc[:,
                        filtered_df_import.columns.get_level_values(0).str.contains("Scope 1, 2, and 3", na=False)]
    df_emission_scope_import = filtered_df_import.loc[:,
                        filtered_df_import.columns.get_level_values(0).str.contains("Scope 1 and 2", na=False)]
    df_emission_limit_import = pd.concat([filtered_df_import[[('', 'Carrier'), ('', 'Unit')]], df_emission_limit_import], axis=1)
    df_emission_scope_import = pd.concat([filtered_df_import[[('', 'Carrier'), ('', 'Unit')]], df_emission_scope_import], axis=1)

    # Ensure output folder exists and save the filtered data
    output_dir = "C:/Users/5637635/PycharmProjects/AdOpT-NET0_Julia/Plotting/Latex"
    os.makedirs(output_dir, exist_ok=True)  # Ensure output folder exists

    filtered_excel_path = os.path.join(output_dir, "filtered_data.xlsx")
    filtered_df = filtered_df.reset_index(drop=True)
    filtered_df.to_excel(filtered_excel_path, index=True, merge_cells=True)

    # Convert the filtered DataFrame to a LaTeX table
    latex_table_limit = (
        "\\begin{table}[h!]\n"
        "\\centering\n"
        "\\caption{Installed capacities for Greenfield, Brownfield and Brownfield with Tight Emission Limit scenarios including Scope 1, 2, and 3 emissions for "
        "the short, mid, and long-term interval}\n"
        "\\label{tab:results_emission_limit}\n"
        "\\begin{adjustbox}{angle=90, max width=\\textheight}"
        + df_emission_limit.to_latex(index=False, escape=False, column_format="lcccccccccccccccc").replace(
        r'\multicolumn{3}{r}{', r'\multicolumn{3}{c}{')
        + "\end{adjustbox}"
        + "\\end{table}"
    )

    latex_table_scope = (
        "\\begin{table}[h!]\n"
        "\\centering\n"
        "\\caption{Installed capacities for Greenfield and Brownfield scenarios including Scope 1 and 2 emissions for "
        "the short, mid, and long-term interval}\n"
        "\\label{tab:results_emission_scope}\n"
        "\\begin{adjustbox}{angle=90, max width=\\textheight}"
        + df_emission_scope.to_latex(index=False, escape=False, column_format="lccccccccccccc").replace(r'\multicolumn{3}{r}{', r'\multicolumn{3}{c}{')
        + "\end{adjustbox}"
        + "\\end{table}"
    )

    #save latex tables
    latex_file_path_limit = os.path.join(output_dir, "filtered_data_emission_limit.tex")
    with open(latex_file_path_limit, "w") as f:
        f.write(latex_table_limit)

    latex_file_path_scope = os.path.join(output_dir, "filtered_data_emission_scope.tex")
    with open(latex_file_path_scope, "w") as f:
        f.write(latex_table_scope)

    #combine import table
    def add_label_row(df, label):
        label_row = pd.DataFrame([['\\textbf{' + label + '}'] + [''] * (df.shape[1] - 1)], columns=df.columns)
        return pd.concat([label_row, df], ignore_index=True)


    #Change header names
    df_emission_limit_import.columns = pd.MultiIndex.from_tuples([
        ('Brownfield' if 'Brownfield' in col[0] else
         'Greenfield' if 'Greenfield' in col[0] else col[0], col[1])
        for col in df_emission_limit_import.columns
    ])
    df_emission_scope_import.columns = pd.MultiIndex.from_tuples([
        ('Brownfield' if 'Brownfield' in col[0] else
         'Greenfield' if 'Greenfield' in col[0] else col[0], col[1])
        for col in df_emission_scope_import.columns
    ])

    #Include import tight
    import_tight = pd.read_pickle("import_tight.pkl")
    import_tight.columns = pd.MultiIndex.from_tuples([
        ('Brownfield' if 'Brownfield' in col[0] else col[0], col[1])
        for col in import_tight.columns
    ])

    # Combine with labels and midrules
    import_combined = pd.concat([
        add_label_row(df_emission_limit_import, "Scope 1, 2, and 3"),
        pd.DataFrame([['\\midrule'] + [''] * (df_emission_limit_import.shape[1] - 1)], columns=df_emission_limit_import.columns),
        add_label_row(import_tight, "Tight Emission Limit (Scope 1, 2, and 3)"),
        pd.DataFrame([['\\midrule'] + [''] * (df_emission_limit_import.shape[1] - 1)],
                     columns=df_emission_limit_import.columns),
        add_label_row(df_emission_scope_import, "Scope 1 and 2")
    ], ignore_index=True)

    import_combined = import_combined.apply(custom_format_row, axis=1)

    # Save to LaTeX
    latex_table_import_combined = (
            "\\begin{table}[h!]\n"
            "\\centering\n"
            "\\caption{Maximum import and export capacities for the Greenfield and Brownfield scenarios, including and excluding Scope 3 emissions, for "
            "the short, mid, and long-term intervals}\n"
            "\\label{tab:results_sensitivity_import_combined}\n"
            "\\begin{adjustbox}{angle=90, max width=\\textheight}\n"
            + import_combined.to_latex(
        index=False,
        escape=False,
        column_format="l" + "c" * (import_combined.shape[1] - 1)
    ).replace(r'\textbackslash midrule', r'\midrule')  # Fix midrule rendering
            + "\\end{adjustbox}\n"
              "\\end{table}"
    )

    with open(os.path.join(output_dir, "filtered_data_import_combined.tex"), "w") as f:
        f.write(latex_table_import_combined)

    import_combined_excel_path = os.path.join(output_dir, "filtered_data_import_combined.xlsx")
    df_import_combined = import_combined.reset_index(drop=True)
    df_import_combined.to_excel(import_combined_excel_path, index=True, merge_cells=True)

#latex tables sensitivity
execute = 0

if execute:
    # Load the Excel file
    file_path = "C:/Users/5637635/PycharmProjects/AdOpT-NET0_Julia/Plotting/result_data_long_sensitivity.xlsx"
    df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)

    # Use the first and third rows as headers
    top_header = df.iloc[0].ffill()  # Forward fill merged cells
    top_header = top_header.replace({"EmissionLimit Greenfield": "Greenfield (Scope 1, 2, and 3)",
                                     "EmissionLimit Brownfield": "Brownfield (Scope 1, 2, and 3)"})
    mid_header = df.iloc[1].ffill()  # Forward fill merged cells
    mid_header = mid_header.replace({"noCO2electrolysis": "No \ce{CO2} electrolyzers",
                                     "MPWemission": "Direct Emissions from MPW Gasification",
                                     "OptBIO": "Optimistic Bio-Feedstock Prices",
                                     "TightEmission": "Tighter Short-Term Emission Limit"})
    sub_header = df.iloc[2].replace({"2030": "Short-term", "2040": "Mid-term", "2050": "Long-term"})

    df.columns = pd.MultiIndex.from_arrays([top_header, mid_header, sub_header])

    # Rename first column header
    top_header.iloc[0] = ""
    mid_header.iloc[0] = ""
    sub_header.iloc[0] = "Result type"
    df = df.iloc[4:].reset_index(drop=True)

    # Define the filter criteria
    keywords = [
        "size_"
    ]
    keywords_import = [
        "electricity/import_max", "CO2/export_max",
        "methane/import_max", "methane_bio/import_max",
        "CO2_DAC/import_max", "MPW/import_max", "propane/import_max"
    ]
    # keywords = [
    #     "size_", "electricity/import_max", "CO2/export_max",
    #     "methane/import_max", "methane_bio/import_max",
    #     "CO2_DAC/import_max", "MPW/import_max", "propane/import_max"
    # ]

    # Extract first column name (handle MultiIndex properly)
    first_column_name = df.columns[0]  # First column from MultiIndex
    first_column_data = df[first_column_name].squeeze()

    # Filter rows based on keywords
    filtered_df = df[first_column_data.astype(str).str.startswith(tuple(keywords))]
    filtered_df_import = df[first_column_data.astype(str).str.startswith(tuple(keywords_import))]

    # Remove unwanted rows
    filtered_df = filtered_df[
        ~first_column_data.str.contains("mixer|CO2toEmission|WGS|OlefinSeparation", na=False, case=False)]

    name_unit_mapping = {
        "size_CrackerFurnace": (r"Conventional cracker", "t naphtha/h"),
        "size_CrackerFurnace_CC": (r"Conventional cracker with \acs{CC}", "\% captured"),
        "size_CrackerFurnace_Electric": (r"Electric cracker", "t naphtha/h"),
        "size_SteamReformer": (r"Conventional reformer", "MW gas"),
        "size_SteamReformer_CC": (r"Conventional reformer with \acs{CC}", "\% captured"),
        "size_ElectricSMR_m": (r"Electric reformer", "MW gas"),
        "size_AEC": (r"\acs{AEC}", "MW electric"),
        "size_HaberBosch": (r"\acs{HB} process", "MW hydrogen"),
        "size_RWGS": ("RWGS", "t \ce{CO2}/h"),
        "size_MeOHsynthesis": (r"Methanol synthesis from syngas", "t syngas/h"),
        "size_DirectMeOHsynthesis": (r"Direct methanol synthesis from \ce{CO2}", "t \ce{CO2}/h"),
        "size_MTO": (r"\acs{MTO}", "t methanol/h"),
        "size_EDH": (r"\acs{EDH}", "t ethanol/h"),
        "size_PDH": (r"\acs{PDH}", "t propane/h"),
        "size_MPW2methanol": (r"\acs{MPW}-to-methanol", "t MPW/h"),
        "size_MPW2methanol_CC": (r"\acs{MPW}-to-methanol with \acs{CC}", "\% captured"),
        "size_CO2electrolysis": (r"\ce{CO2} electrolysis", "t \ce{CO2}/h"),
        "size_ASU": (r"\acs{ASU}", "MW electricity"),
        "size_Boiler_Industrial_NG": (r"Gas-fired boiler", "MW gas"),
        "size_Boiler_El": (r"Electric boiler", "MW electricity"),
        # Storage & Import Entries (Keep These at the Bottom)
        "size_Storage_Ammonia": ("Ammonia tank", "tonne"),
        "size_Storage_Battery": ("Li-ion battery", "MWh"),
        "size_Storage_CO2": (r"\ce{CO2} buffer storage", "tonne"),
        "size_Storage_H2": ("Hydrogen tank", "MWh"),
        "size_Storage_Ethylene": ("Ethylene tank", "tonne"),
        "size_Storage_Propylene": ("Propylene tank", "tonne"),
        # "electricity/import_max": ("Electricity grid import", "MW (\% of max)"),
        # "CO2/export_max": (r"\ce{CO2} T\&S", "t \ce{CO2}/h (\% of max)"),
        # "MPW/import_max": ("MPW import", "t MPW/h (\% of max)"),
        # "methane/import_max": ("Methane import", "MW gas"),
        # "methane_bio/import_max": ("Bio-methane import", "MW gas"),
        # "CO2_DAC/import_max": ("DAC-\ce{CO2} import", "t \ce{CO2}/h"),
        # "propane/import_max": ("Bio-propane import", "MW"),
    }

    name_unit_mapping_import = {
        "electricity/import_max": ("Electricity grid import", "MW (\% of max)"),
        "CO2/export_max": (r"\ce{CO2} T\&S", "t \ce{CO2}/h (\% of max)"),
        "MPW/import_max": ("MPW import", "t MPW/h (\% of max)"),
        "methane/import_max": ("Methane import", "MW gas"),
        "methane_bio/import_max": ("Bio-methane import", "MW gas"),
        "CO2_DAC/import_max": ("DAC-\ce{CO2} import", "t \ce{CO2}/h"),
        "propane/import_max": ("Bio-propane import", "MW"),
    }

    # Function to rename technologies and format existing ones correctly
    def rename_tech(name):
        base_name = name.replace("_existing", "")
        tech_name, unit = name_unit_mapping.get(base_name, (name, "-"))  # Default to '-' if not found
        if "_existing" in name:
            tech_name += " existing"
        return tech_name, unit

    def rename_tech_import(name):
        tech_name, unit = name_unit_mapping_import.get(name, (name, "-"))  # Default to '-' if not found
        return tech_name, unit


    # Apply renaming and extract units using .map() for better performance
    filtered_df[('', '', 'Technology')] = filtered_df[first_column_name].map(lambda x: rename_tech(x)[0])
    filtered_df[('', '', 'Unit')] = filtered_df[first_column_name].map(lambda x: rename_tech(x)[1])
    filtered_df_import[('', '', 'Carrier')] = filtered_df_import[first_column_name].map(lambda x: rename_tech_import(x)[0])
    filtered_df_import[('', '', 'Unit')] = filtered_df_import[first_column_name].map(lambda x: rename_tech_import(x)[1])

    # Define order based on LaTeX table
    ordered_techs = list(name_unit_mapping.keys())
    ordered_imports = list(name_unit_mapping_import.keys())

    # Generate a new list of technologies with "existing" tech placed right below the original technology
    ordered_techs_with_existing = []
    for tech in ordered_techs:
        if "CC" not in tech:
            ordered_techs_with_existing.append(tech)
            if f"{tech}_CC" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_CC")
            if f"{tech}_existing" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_existing")
            if f"{tech}_existing_CC" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_existing_CC")

    # Create sorting index to maintain the correct order (with "existing" techs below the original ones)
    filtered_df['Order'] = first_column_data.apply(
        lambda x: ordered_techs_with_existing.index(x) if x in ordered_techs_with_existing else len(
            ordered_techs_with_existing))
    filtered_df_import['Order'] = first_column_data.apply(
        lambda x: ordered_imports.index(x) if x in name_unit_mapping_import else len(
            name_unit_mapping_import))

    # Sort the dataframe based on the 'Order' column and remove it
    filtered_df = filtered_df.sort_values(by=['Order']).drop(columns=['Order'])
    filtered_df_import = filtered_df_import.sort_values(by=['Order']).drop(columns=['Order'])

    # Round all numeric columns to whole numbers and format nan
    def custom_format_row(row):
        if 'Resulttype' in row and 'CC' in str(row['Resulttype']):
            # Preserve floats (just replace NaN with '-')
            return row.apply(lambda x: '-' if pd.isna(x) else int(x * 100) if isinstance(x, float) else x)
        else:
            # Round floats (except NaN) to int
            return row.apply(lambda x: '-' if pd.isna(x) else int(round(x, 0)) if isinstance(x, float) else x)

    filtered_df = filtered_df.apply(custom_format_row, axis=1)
    filtered_df_import = filtered_df_import.apply(custom_format_row, axis=1)

    # If the index is multi-level, you can access the levels
    ordered_columns = [('', '', 'Technology'), ('', '', 'Unit')] + [col for col in filtered_df.columns if
                                                            col not in [('', '', 'Technology'), ('', '', 'Unit')]]
    ordered_columns_import = [('', '', 'Carrier'), ('', '', 'Unit')] + [col for col in filtered_df_import.columns if
                                                                    col not in [('', '', 'Carrier'),
                                                                                ('', '', 'Unit')]]
    filtered_df = filtered_df[ordered_columns]
    filtered_df = filtered_df.drop(columns=first_column_name)
    filtered_df = filtered_df.drop(columns=[('Greenfield (Scope 1, 2, and 3)', 'Tighter Short-Term Emission Limit')])
    filtered_df_import = filtered_df_import[ordered_columns_import]
    filtered_df_import = filtered_df_import.drop(columns=first_column_name)
    filtered_df_import = filtered_df_import.drop(columns=[('Greenfield (Scope 1, 2, and 3)', 'Tighter Short-Term Emission Limit')])

    def swap_cc_existing(tech_str):
        if isinstance(tech_str, str):
            return re.sub(r'with (\\acs\{CC\}) existing', r'existing with \1', tech_str)
        return tech_str

    filtered_df[('', '', 'Technology')] = filtered_df[('', '', 'Technology')].apply(swap_cc_existing)

    # Ensure output folder exists and save the filtered data
    output_dir = "C:/Users/5637635/PycharmProjects/AdOpT-NET0_Julia/Plotting/Latex"
    os.makedirs(output_dir, exist_ok=True)  # Ensure output folder exists

    # Split filtered_df into three based on the mid-header
    mid_levels = [
        'No \ce{CO2} electrolyzers'
        'Direct Emissions from MPW Gasification',
        'Optimistic Bio-Feedstock Prices',
        'Tighter Short-Term Emission Limit'
    ]

    # Function to filter by mid-header value
    def split_by_mid_header(df, mid_value):
        return df.loc[:, df.columns.get_level_values(1) == mid_value]


    # Function to drop the mid-header level from MultiIndex
    def drop_mid_header(df):
        df.columns = pd.MultiIndex.from_arrays([
            df.columns.get_level_values(0),
            df.columns.get_level_values(2)
        ])
        return df

    # Keep first two columns (technology/unit) from original df
    base_cols = filtered_df.iloc[:, :2]  # Assuming these are 'Technology' and 'Unit'

    # Process each subset
    df_noCO2_raw = split_by_mid_header(filtered_df, 'No \ce{CO2} electrolyzers')
    df_mpw_raw = split_by_mid_header(filtered_df, 'Direct Emissions from MPW Gasification')
    df_optbio_raw = split_by_mid_header(filtered_df, 'Optimistic Bio-Feedstock Prices')
    df_tight_raw = split_by_mid_header(filtered_df, 'Tighter Short-Term Emission Limit')

    # Clean columns
    df_noCO2 = drop_mid_header(pd.concat([base_cols, df_noCO2_raw], axis=1))
    df_mpw = drop_mid_header(pd.concat([base_cols, df_mpw_raw], axis=1))
    df_optbio = drop_mid_header(pd.concat([base_cols, df_optbio_raw], axis=1))
    df_tight = drop_mid_header(pd.concat([base_cols, df_tight_raw], axis=1))  # No tech/unit columns here

    #Reset columns
    df_noCO2 = df_noCO2.reset_index(drop=True)
    df_mpw = df_mpw.reset_index(drop=True)
    df_optbio = df_optbio.reset_index(drop=True)
    df_tight = df_tight.reset_index(drop=True)

    #save to pickle
    df_tight.to_pickle("df_tight.pkl")


    # Optional: define column format length dynamically
    ncols = len(df_mpw.columns)
    column_format = "ll" + "c" * (ncols - 2)  # 'l' for tech/carrier + unit, rest are numeric

    # Create LaTeX tables
    def make_latex(df, caption, label):
        return (
                "\\begin{table}[h!]\n"
                "\\centering\n"
                f"\\caption{{{caption}}}\n"
                f"\\label{{{label}}}\n"
                "\\begin{adjustbox}{angle=90, max width=\\textheight}\n"
                + df.to_latex(index=False, multicolumn=True, multicolumn_format='c', escape=False, column_format=column_format).replace(
            r'\multicolumn{2}{r}{', r'\multicolumn{2}{c}{')
                + "\\end{adjustbox}\n"
                  "\\end{table}"
        )


    # Create individual LaTeX tables
    latex_table_noCO2 = make_latex(df_noCO2,
                                 "Installed capacities for sensitivity case: No \ce{CO2} electrolyzers",
                                 "tab:results_sensitivity_noCO2")
    latex_table_mpw = make_latex(df_mpw,
                                 "Installed capacities for sensitivity case: Direct Emissions from MPW Gasification",
                                 "tab:results_sensitivity_mpw")
    latex_table_optbio = make_latex(df_optbio,
                                    "Installed capacities for sensitivity case: Optimistic Bio-Feedstock Prices",
                                    "tab:results_sensitivity_optbio")
    latex_table_tight = make_latex(df_tight,
                                   "Installed capacities for sensitivity case: Tighter Short-Term Emission Limit",
                                   "tab:results_sensitivity_tight")

    # Save to .tex
    with open(os.path.join(output_dir, "filtered_data_sensitivity_noCO2.tex"), "w") as f:
        f.write(latex_table_noCO2)
    with open(os.path.join(output_dir, "filtered_data_sensitivity_mpw.tex"), "w") as f:
        f.write(latex_table_mpw)
    with open(os.path.join(output_dir, "filtered_data_sensitivity_optbio.tex"), "w") as f:
        f.write(latex_table_optbio)

    #save to excel
    filtered_excel_path = os.path.join(output_dir, "filtered_data_sensitivity.xlsx")
    filtered_df = filtered_df.reset_index(drop=True)
    filtered_df.to_excel(filtered_excel_path, index=True, merge_cells=True)
    filtered_excel_path_import = os.path.join(output_dir, "filtered_data_sensitivity_import.xlsx")
    filtered_df_import = filtered_df_import.reset_index(drop=True)
    filtered_df_import.to_excel(filtered_excel_path_import, index=True, merge_cells=True)

    # === EXPORT FILTERED IMPORT DATA AS STACKED TABLE FOR ALL SENSITIVITIES ===
    # add base cols
    base_cols_import = filtered_df_import.iloc[:, :2]

    # Split imports per case
    import_noCO2_raw = split_by_mid_header(filtered_df_import, 'No \ce{CO2} electrolyzers')
    import_mpw_raw = split_by_mid_header(filtered_df_import, 'Direct Emissions from MPW Gasification')
    import_optbio_raw = split_by_mid_header(filtered_df_import, 'Optimistic Bio-Feedstock Prices')
    import_tight_raw = split_by_mid_header(filtered_df_import, 'Tighter Short-Term Emission Limit')

    import_noCO2 = drop_mid_header(pd.concat([base_cols_import, import_noCO2_raw], axis=1))
    import_mpw = drop_mid_header(pd.concat([base_cols_import, import_mpw_raw], axis=1))
    import_optbio = drop_mid_header(pd.concat([base_cols_import, import_optbio_raw], axis=1))
    import_tight = drop_mid_header(pd.concat([base_cols_import, import_tight_raw], axis=1))

    #save tight to pickle
    import_tight.to_pickle("import_tight.pkl")

    def add_label_row(df, label):
        label_row = pd.DataFrame([['\\textbf{' + label + '}'] + [''] * (df.shape[1] - 1)], columns=df.columns)
        return pd.concat([label_row, df], ignore_index=True)

    # Include import tight
    import_zeeland = pd.read_pickle("df_zeeland_import.pkl")
    # import_zeeland.columns = pd.MultiIndex.from_tuples([
    #     ('Brownfield' if 'Brownfield' in col[0] else col[0], col[1])
    #     for col in import_zeeland.columns
    # ])

    # Combine with labels and midrules
    import_combined = pd.concat([
        add_label_row(import_zeeland, "Zeeland"),
        pd.DataFrame([['\\midrule'] + [''] * (import_mpw.shape[1] - 1)], columns=import_mpw.columns),
        add_label_row(import_noCO2, "No \ce{CO2} electrolyzers"),
        pd.DataFrame([['\\midrule'] + [''] * (import_mpw.shape[1] - 1)], columns=import_mpw.columns),
        add_label_row(import_optbio, "Optimistic Bio-Feedstock Prices"),
        pd.DataFrame([['\\midrule'] + [''] * (import_mpw.shape[1] - 1)], columns=import_mpw.columns),
        add_label_row(import_mpw, "Direct Emissions from MPW Gasification"),
    ], ignore_index=True)



    # Save to LaTeX
    latex_table_import_combined = (
            "\\begin{table}[h!]\n"
            "\\centering\n"
            "\\caption{Maximum import and export capacities for all sensitivity cases}\n"
            "\\label{tab:results_sensitivity_import_combined}\n"
            "\\begin{adjustbox}{angle=90, max width=\\textheight}\n"
            + import_combined.to_latex(
        index=False,
        escape=False,
        column_format="l" + "c" * (import_combined.shape[1] - 1)
    ).replace(r'\textbackslash midrule', r'\midrule')  # Fix midrule rendering
            + "\\end{adjustbox}\n"
              "\\end{table}"
    )

    with open(os.path.join(output_dir, "filtered_data_sensitivity_import_combined.tex"), "w") as f:
        f.write(latex_table_import_combined)


# for zeeland case study
execute = 0

if execute:
    # Load the Excel file
    file_path = "C:/Users/5637635/PycharmProjects/AdOpT-NET0_Julia/Plotting/result_data_long_Zeeland.xlsx"
    df = pd.read_excel(file_path, sheet_name="Sheet1", header=None)

    # Use the first and third rows as headers
    top_header = df.iloc[0].ffill()  # Forward fill merged cells
    top_header = top_header.replace({"EmissionLimit Greenfield": "Greenfield (Scope 1, 2, and 3)", "EmissionLimit Brownfield": "Brownfield (Scope 1, 2, and 3)",
                                     "EmissionScope Greenfield": "Greenfield (Scope 1 and 2)", "EmissionScope Brownfield":
                                         "Brownfield (Scope 1 and 2)",})
    sub_header = df.iloc[2].replace({"2030": "Short-term", "2040": "Mid-term", "2050": "Long-term"})

    # sub_header = df.iloc[2].astype(str)  # Convert to string
    df.columns = pd.MultiIndex.from_arrays([top_header, sub_header])

    # Rename first column header
    top_header = top_header.copy()  # Avoid modifying the original header list
    top_header.iloc[0] = ""
    sub_header.iloc[0] = "Result type"
    df = df.iloc[4:].reset_index(drop=True)

    # Define the filter criteria
    keywords = [
        "size_"
    ]
    keywords_import = [
        "electricity/import_max", "CO2/export_max",
        "methane/import_max", "methane_bio/import_max",
        "CO2_DAC/import_max", "MPW/import_max", "propane/import_max"
    ]

    # Extract first column name
    first_column_name = df.columns[0]
    first_column_data = df[first_column_name].squeeze()

    # Filter rows based on keywords
    filtered_df = df[df[first_column_name].astype(str).str.startswith(tuple(keywords))]
    filtered_df_import = df[first_column_data.astype(str).str.startswith(tuple(keywords_import))]

    # Remove unwanted rows
    filtered_df = filtered_df[
        ~filtered_df[first_column_name].str.contains("mixer|CO2toEmission|WGS|OlefinSeparation", na=False, case=False)]

    name_unit_mapping = {
        "size_CrackerFurnace": (r"Conventional cracker", "t naphtha/h"),
        "size_CrackerFurnace_CC": (r"Conventional cracker with \acs{CC}", "\% captured"),
        "size_CrackerFurnace_Electric": (r"Electric cracker", "t naphtha/h"),
        "size_SteamReformer": (r"Conventional reformer", "MW gas"),
        "size_SteamReformer_CC": (r"Conventional reformer with \acs{CC}", "\% captured"),
        "size_ElectricSMR_m": (r"Electric reformer", "MW gas"),
        "size_AEC": (r"\acs{AEC}", "MW electric"),
        "size_HaberBosch": (r"\acs{HB} process", "MW hydrogen"),
        "size_RWGS": ("RWGS", "t \ce{CO2}/h"),
        "size_MeOHsynthesis": (r"Methanol synthesis from syngas", "t syngas/h"),
        "size_DirectMeOHsynthesis": (r"Direct methanol synthesis from \ce{CO2}", "t \ce{CO2}/h"),
        "size_MTO": (r"\acs{MTO}", "t methanol/h"),
        "size_EDH": (r"\acs{EDH}", "t ethanol/h"),
        "size_PDH": (r"\acs{PDH}", "t propane/h"),
        "size_MPW2methanol": (r"\acs{MPW}-to-methanol", "t MPW/h"),
        "size_MPW2methanol_CC": (r"\acs{MPW}-to-methanol with \acs{CC}", "\% captured"),
        "size_CO2electrolysis": (r"\ce{CO2} electrolysis", "t \ce{CO2}/h"),
        "size_ASU": (r"\acs{ASU}", "MW electricity"),
        "size_Boiler_Industrial_NG": (r"Gas-fired boiler", "MW gas"),
        "size_Boiler_El": (r"Electric boiler", "MW electricity"),
        # Storage & Import Entries (Keep These at the Bottom)
        "size_Storage_Ammonia": ("Ammonia tank", "tonne"),
        "size_Storage_Battery": ("Li-ion battery", "MWh"),
        "size_Storage_CO2": (r"\ce{CO2} buffer storage", "tonne"),
        "size_Storage_H2": ("Hydrogen tank", "MWh"),
        "size_Storage_Ethylene": ("Ethylene tank", "tonne"),
        "size_Storage_Propylene": ("Propylene tank", "tonne"),
        # "electricity/import_max": ("Electricity grid import", "MW (\% of max)"),
        # "CO2/export_max": (r"\ce{CO2} T\&S", "t \ce{CO2}/h (\% of max)"),
        # "MPW/import_max": ("MPW import", "t MPW/h (\% of max)"),
        # "methane/import_max": ("Methane import", "MW gas"),
        # "methane_bio/import_max": ("Bio-methane import", "MW gas"),
        # "CO2_DAC/import_max": ("DAC-\ce{CO2} import", "t \ce{CO2}/h"),
        # "propane/import_max": ("Bio-propane import", "MW"),
    }

    name_unit_mapping_import = {
        "electricity/import_max": ("Electricity grid import", "MW (\% of max)"),
        "CO2/export_max": (r"\ce{CO2} T\&S", "t \ce{CO2}/h (\% of max)"),
        "MPW/import_max": ("MPW import", "t MPW/h (\% of max)"),
        "methane/import_max": ("Methane import", "MW gas"),
        "methane_bio/import_max": ("Bio-methane import", "MW gas"),
        "CO2_DAC/import_max": ("DAC-\ce{CO2} import", "t \ce{CO2}/h"),
        "propane/import_max": ("Bio-propane import", "MW"),
    }


    # Function to rename technologies and format existing ones correctly
    def rename_tech(name):
        base_name = name.replace("_existing", "")
        tech_name, unit = name_unit_mapping.get(base_name, (name, "-"))  # Default to '-' if not found
        if "_existing" in name:
            tech_name += " existing"
        return tech_name, unit

    def rename_import(name):
        import_name, unit = name_unit_mapping_import.get(name, (name, "-"))  # Default to '-' if not found
        return import_name, unit

    # Apply renaming and extract units using .map() for better performance
    filtered_df[('', 'Technology')] = filtered_df[first_column_name].map(lambda x: rename_tech(x)[0])
    filtered_df[('', 'Unit')] = filtered_df[first_column_name].map(lambda x: rename_tech(x)[1])
    filtered_df_import[( '', 'Carrier')] = filtered_df_import[first_column_name].map(
        lambda x: rename_import(x)[0])
    filtered_df_import[('', 'Unit')] = filtered_df_import[first_column_name].map(lambda x: rename_import(x)[1])

    # Define order based on LaTeX table
    ordered_techs = list(name_unit_mapping.keys())
    ordered_imports = list(name_unit_mapping_import.keys())

    # Generate a new list of technologies with "existing" tech placed right below the original technology
    ordered_techs_with_existing = []
    for tech in ordered_techs:
        if "CC" not in tech:
            ordered_techs_with_existing.append(tech)
            if f"{tech}_CC" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_CC")
            if f"{tech}_existing" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_existing")
            if f"{tech}_existing_CC" in filtered_df[first_column_name].values:
                ordered_techs_with_existing.append(f"{tech}_existing_CC")

    # Create sorting index to maintain the correct order (with "existing" techs below the original ones)
    filtered_df['Order'] = filtered_df[first_column_name].apply(
        lambda x: ordered_techs_with_existing.index(x) if x in ordered_techs_with_existing else len(
            ordered_techs_with_existing))
    filtered_df_import['Order'] = first_column_data.apply(
        lambda x: ordered_imports.index(x) if x in name_unit_mapping_import else len(
            name_unit_mapping_import))

    # Sort the dataframe based on the 'Order' column and remove it
    filtered_df = filtered_df.sort_values(by=['Order']).drop(columns=['Order'])
    filtered_df_import = filtered_df_import.sort_values(by=['Order']).drop(columns=['Order'])

    # Round all numeric columns to whole numbers and format nan
    def custom_format_row(row):
        if 'Resulttype' in row and 'CC' in str(row['Resulttype']):
            # Preserve floats (just replace NaN with '-')
            return row.apply(lambda x: '-' if pd.isna(x) else int(x*100) if isinstance(x, float) else x)
        else:
            # Round floats (except NaN) to int
            return row.apply(lambda x: '-' if pd.isna(x) else int(round(x, 0)) if isinstance(x, float) else x)


    filtered_df = filtered_df.apply(custom_format_row, axis=1)
    filtered_df_import = filtered_df_import.apply(custom_format_row, axis=1)

    # If the index is multi-level, you can access the levels
    ordered_columns = [('', 'Technology') , ('', 'Unit')] + [col for col in filtered_df.columns if col not in [('', 'Technology'), ('', 'Unit')]]
    ordered_columns_import = [('', 'Carrier') , ('', 'Unit')] + [col for col in filtered_df_import.columns if col not in [('', 'Carrier'), ('', 'Unit')]]
    filtered_df = filtered_df[ordered_columns]
    filtered_df_import = filtered_df_import[ordered_columns_import]
    filtered_df = filtered_df.drop(columns=('Resulttype', 'Interval'))
    filtered_df_import = filtered_df_import.drop(columns=('Resulttype', 'Interval'))

    # save to pickle
    filtered_df_import.to_pickle("df_zeeland_import.pkl")

    def swap_cc_existing(tech_str):
        if isinstance(tech_str, str):
            return re.sub(r'with (\\acs\{CC\}) existing', r'existing with \1', tech_str)
        return tech_str

    # Apply to the relevant column
    filtered_df[('', 'Technology')] = filtered_df[('', 'Technology')].apply(swap_cc_existing)

    # Ensure output folder exists and save the filtered data
    output_dir = "C:/Users/5637635/PycharmProjects/AdOpT-NET0_Julia/Plotting/Latex"
    os.makedirs(output_dir, exist_ok=True)  # Ensure output folder exists

    filtered_excel_path = os.path.join(output_dir, "filtered_data_Zeeland.xlsx")
    filtered_df = filtered_df.reset_index(drop=True)
    filtered_df.to_excel(filtered_excel_path, index=True, merge_cells=True)

    # Convert the filtered DataFrame to a LaTeX table
    latex_table_zeeland = (
            "\\begin{table}[h!]\n"
            "\\centering\n"
            "\\caption{Installed capacities for sensitivity case: Zeeland}\n"
            "\\label{tab:results_emission_limit}\n"
            "\\begin{adjustbox}{angle=90, max width=\\textheight}"
            + filtered_df.to_latex(index=False, escape=False, column_format="llcccccc").replace(
        r'\multicolumn{3}{r}{', r'\multicolumn{3}{c}{')
            + "\end{adjustbox}"
            + "\\end{table}"
    )

    with open(os.path.join(output_dir, "filtered_data_zeeland.tex"), "w") as f:
        f.write(latex_table_zeeland)
